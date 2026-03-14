from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from app.model.lib.excel import export_to_xlsx

_TIME_UNITS = {
    'd': 'days',
    'h': 'hours',
    'm': 'minutes',
    's': 'seconds',
}

_TECHNIQUE_DESCRIPTIONS = {
    'fc':          "Flow-cytometry values per time-point in {units}, use a period (.) as the decimal separator",
    'fc_ps':       "FC counts values per time-point for strain {strain} in {units}, use a period (.) as the decimal separator",
    'od':          "Optical density values per time-point, use a period (.) as the decimal separator",
    'plates':      "Plate count values per time-point, use a period (.) as the decimal separator",
    'plates_ps':   "Plate count values per time-point for strain {strain}, use a period (.) as the decimal separator.",
    '16s_ps':      "Abundances values per time-point for strain {strain}, use a period (.) as the decimal separator.",
    'qpcr':        "Abundances values per time-point, use a period (.) as the decimal separator.",
    'qpcr_ps':     "Abundances values per time-point for strain {strain}, use a period (.) as the decimal separator.",
    'ph':          "pH values per time-point, use a period (.) as the decimal separator",
    'metabolites': "Concentration values per time-point for metabolite {metabolite} in {units}, use a period (.) as the decimal separator",
    'STD':         "Standard deviation if more than 1 technical replicate, use a period (.) as the decimal separator",
}


def create_excel(submission_form):
    "Create a template data file based on the submission's study design"

    submission = submission_form.submission

    strain_names = [t.name for t in submission_form.fetch_taxa()]
    strain_names += [s['name'] for s in submission.studyDesign['custom_strains']]

    workbook = Workbook()

    short_time_units = submission.studyDesign['timeUnits']
    long_time_units = _TIME_UNITS[short_time_units]

    headers_common = {
        'Biological Replicate': 'Unique names of individual samples: a bottle, a well in a well-plate, or a mini-bioreactor.',
        'Compartment':          'A compartment within the vessel where growth is measured.',
        'Time':                 f"Measurement time-points in {long_time_units} ({short_time_units}).",
    }

    headers_bioreplicates = {**headers_common}
    headers_strains       = {**headers_common}
    headers_metabolites   = {**headers_common}

    for index, study_technique in enumerate(submission.build_techniques()):
        subject_type   = study_technique.subjectType
        technique_type = study_technique.type
        units          = study_technique.units

        for measurement_technique in study_technique.measurementTechniques:
            if subject_type == 'bioreplicate':
                technique_name = measurement_technique.csv_column_name()
                description = _TECHNIQUE_DESCRIPTIONS[technique_type].format(units=study_technique.units)

                headers_bioreplicates[technique_name] = description

                if study_technique.includeStd:
                    title = ' '.join([technique_name, 'STD'])
                    headers_bioreplicates[title] = _TECHNIQUE_DESCRIPTIONS['STD']

            elif subject_type == 'strain':
                if study_technique.includeUnknown:
                    technique_strain_names = [*strain_names, "Unknown"]
                else:
                    technique_strain_names = strain_names

                for strain_name in technique_strain_names:
                    title = measurement_technique.csv_column_name(strain_name)

                    description = _TECHNIQUE_DESCRIPTIONS[f"{technique_type}_ps"].format(
                        strain=strain_name,
                        units=units,
                    )
                    headers_strains[title] = description

                    if study_technique.includeStd:
                        title = ' '.join([title, 'STD'])
                        headers_strains[title] = _TECHNIQUE_DESCRIPTIONS['STD']

            elif subject_type == 'metabolite':
                metabolites = submission_form.fetch_metabolites_for_technique(index)

                for metabolite in metabolites:
                    title = measurement_technique.csv_column_name(metabolite.name)
                    description = _TECHNIQUE_DESCRIPTIONS['metabolites'].format(
                        metabolite=metabolite.name,
                        units=units,
                    )
                    headers_metabolites[title] = description

                    if study_technique.includeStd:
                        title = ' '.join([metabolite.name, 'STD'])
                        headers_metabolites[title] = _TECHNIQUE_DESCRIPTIONS['STD']

            else:
                raise ValueError(f"Invalid technique subject_type: {subject_type}")

    # Create sheets for each category of measurement:
    if len(headers_bioreplicates) > 3:
        _fill_sheet(workbook, "Growth data per community", headers_bioreplicates, submission)

    if len(headers_strains) > 3:
        _fill_sheet(workbook, "Growth data per strain", headers_strains, submission)

    if len(headers_metabolites) > 3:
        _fill_sheet(workbook, "Growth data per metabolite", headers_metabolites, submission)

    return export_to_xlsx(workbook)


def _add_header(sheet, index, title, description):
    cell         = sheet.cell(row=1, column=index, value=title)
    cell.comment = Comment(description, author="μGrowthDB")

    # Built-in styles:
    # <https://openpyxl.readthedocs.io/en/stable/styles.html#using-builtin-styles>
    #
    cell.style = 'Headline 3'

    # Calculate column width based on the length of the title text, plus some extra space
    column_width = len(title)
    column_letter = get_column_letter(index)
    sheet.column_dimensions[column_letter].width = column_width + 1


def _fill_sheet(workbook, sheet_title, headers, submission):
    if workbook.sheetnames == ['Sheet']:
        # Then we have a brand new workbook, use its first sheet
        sheet = workbook.active
        workbook.active.title = sheet_title
    else:
        sheet = workbook.create_sheet(title=sheet_title)

    # Add headers and descriptions to the first row and modify the width of each columns
    for index, (title, description) in enumerate(headers.items(), start=1):
        _add_header(sheet, index, title, description)

    bottom_border = Border(bottom=Side(style="thin", color="000000"))

    row_index = 2
    for experiment in submission.studyDesign['experiments']:
        for bioreplicate in experiment['bioreplicates']:
            for compartment_name in experiment['compartmentNames']:
                for _ in range(experiment['timepointCount']):
                    sheet.cell(row=row_index, column=1, value=bioreplicate['name'])
                    sheet.cell(row=row_index, column=2, value=compartment_name)

                    row_index += 1

                # Apply border after every group of time points
                for column in range(1, len(headers.keys()) + 1):
                    cell = sheet.cell(row=(row_index - 1), column=column)
                    cell.border = bottom_border

    # Freeze header and label columns
    sheet.freeze_panes = "D2"
